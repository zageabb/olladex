from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


_TABLE_NAME = re.compile(r"^[A-Za-z_][A-Za-z0-9_.]*$")


def inspect_xlsx(path: Path) -> dict:
    book = load_workbook(path, read_only=False, data_only=False)
    try:
        sheets = []
        for sheet in book.worksheets:
            max_row = min(max(sheet.max_row, 1), 200)
            max_column = min(max(sheet.max_column, 1), 100)
            rows = []
            cells = []
            formulas = []
            for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_column):
                values = []
                for cell in row:
                    values.append(cell.value)
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formulas.append({"cell": cell.coordinate, "formula": cell.value})
                    if cell.value is not None or cell.has_style:
                        fill_color = cell.fill.fgColor.rgb if cell.fill.fill_type else ""
                        font_color = cell.font.color.rgb if cell.font.color and cell.font.color.type == "rgb" else ""
                        cells.append(
                            {
                                "cell": cell.coordinate,
                                "value": cell.value,
                                "formula": cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else "",
                                "number_format": cell.number_format,
                                "bold": bool(cell.font.bold),
                                "italic": bool(cell.font.italic),
                                "font_size": cell.font.sz,
                                "font_color": str(font_color or ""),
                                "fill_color": str(fill_color or ""),
                                "horizontal": cell.alignment.horizontal or "",
                                "vertical": cell.alignment.vertical or "",
                                "wrap_text": bool(cell.alignment.wrap_text),
                            }
                        )
                rows.append(values)
            sheets.append(
                {
                    "name": sheet.title,
                    "rows": rows,
                    "cells": cells,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "formulas": formulas[:1000],
                    "merged_ranges": [str(value) for value in sheet.merged_cells.ranges],
                    "freeze_panes": str(sheet.freeze_panes or ""),
                    "auto_filter": sheet.auto_filter.ref or "",
                    "tables": [
                        {
                            "name": table.name,
                            "display_name": table.displayName,
                            "ref": table.ref,
                            "style": table.tableStyleInfo.name if table.tableStyleInfo else "",
                        }
                        for table in sheet.tables.values()
                    ],
                    "hidden": sheet.sheet_state != "visible",
                }
            )
        return {"kind": "excel", "sheets": sheets, "defined_names": [name for name in book.defined_names]}
    finally:
        book.close()


def mutate_xlsx(path: Path, operations: list[dict[str, Any]]) -> None:
    book = load_workbook(path, read_only=False, data_only=False)
    try:
        for operation in operations:
            action = operation.get("action")
            if action == "set_cell":
                sheet = _sheet(book, operation.get("sheet"))
                cell = _cell(operation)
                sheet[cell] = operation.get("value")
            elif action == "set_cell_format":
                sheet = _sheet(book, operation.get("sheet"))
                cell = sheet[_cell(operation)]
                font_color = _hex(operation.get("font_color")) if "font_color" in operation else _font_rgb(cell)
                cell.font = Font(
                    name=str(operation.get("font_name") or cell.font.name or "Calibri"),
                    size=float(operation.get("font_size") or cell.font.sz or 11),
                    bold=bool(operation.get("bold")) if "bold" in operation else bool(cell.font.bold),
                    italic=bool(operation.get("italic")) if "italic" in operation else bool(cell.font.italic),
                    underline="single" if operation.get("underline") else (cell.font.underline if "underline" not in operation else None),
                    color=font_color or None,
                )
                if "fill_color" in operation:
                    fill = _hex(operation.get("fill_color"))
                    cell.fill = PatternFill(fill_type="solid", fgColor=fill) if fill else PatternFill()
                cell.alignment = Alignment(
                    horizontal=str(operation.get("horizontal") or cell.alignment.horizontal or "general"),
                    vertical=str(operation.get("vertical") or cell.alignment.vertical or "bottom"),
                    wrap_text=bool(operation.get("wrap_text")) if "wrap_text" in operation else bool(cell.alignment.wrap_text),
                )
                if "number_format" in operation:
                    cell.number_format = str(operation.get("number_format") or "General")
            elif action == "set_range_values":
                sheet = _sheet(book, operation.get("sheet"))
                start = _cell(operation, "start_cell")
                anchor = sheet[start]
                values = operation.get("values")
                if not isinstance(values, list):
                    raise ValueError("set_range_values requires values")
                for row_offset, row in enumerate(values):
                    if not isinstance(row, list):
                        continue
                    for column_offset, value in enumerate(row):
                        sheet.cell(row=anchor.row + row_offset, column=anchor.column + column_offset, value=value)
            elif action == "add_sheet":
                name = _sheet_name(operation.get("name"))
                if name in book.sheetnames:
                    raise ValueError(f"Worksheet already exists: {name}")
                book.create_sheet(title=name)
            elif action == "rename_sheet":
                sheet = _sheet(book, operation.get("sheet"))
                name = _sheet_name(operation.get("name"))
                if name in book.sheetnames and name != sheet.title:
                    raise ValueError(f"Worksheet already exists: {name}")
                sheet.title = name
            elif action == "delete_sheet":
                if len(book.worksheets) <= 1:
                    raise ValueError("Workbook must contain at least one worksheet")
                book.remove(_sheet(book, operation.get("sheet")))
            elif action in {"insert_rows", "delete_rows"}:
                sheet = _sheet(book, operation.get("sheet"))
                index = max(1, int(operation.get("index", 1)))
                amount = max(1, int(operation.get("amount", 1)))
                sheet.insert_rows(index, amount) if action == "insert_rows" else sheet.delete_rows(index, amount)
            elif action in {"insert_columns", "delete_columns"}:
                sheet = _sheet(book, operation.get("sheet"))
                index = max(1, int(operation.get("index", 1)))
                amount = max(1, int(operation.get("amount", 1)))
                sheet.insert_cols(index, amount) if action == "insert_columns" else sheet.delete_cols(index, amount)
            elif action == "merge_cells":
                _sheet(book, operation.get("sheet")).merge_cells(str(operation.get("range", "")).strip())
            elif action == "unmerge_cells":
                _sheet(book, operation.get("sheet")).unmerge_cells(str(operation.get("range", "")).strip())
            elif action == "freeze_panes":
                sheet = _sheet(book, operation.get("sheet"))
                value = str(operation.get("cell", "")).strip().upper()
                sheet.freeze_panes = value or None
            elif action == "set_auto_filter":
                sheet = _sheet(book, operation.get("sheet"))
                sheet.auto_filter.ref = str(operation.get("range", "")).strip() or None
            elif action == "set_column_width":
                sheet = _sheet(book, operation.get("sheet"))
                column = str(operation.get("column", "A")).strip().upper()
                sheet.column_dimensions[column].width = float(operation.get("width", 12))
            elif action == "set_row_height":
                sheet = _sheet(book, operation.get("sheet"))
                row = max(1, int(operation.get("row", 1)))
                sheet.row_dimensions[row].height = float(operation.get("height", 15))
            elif action == "add_table":
                sheet = _sheet(book, operation.get("sheet"))
                ref = str(operation.get("range", "")).strip().upper()
                name = str(operation.get("name", "Table1")).strip()
                if not ref:
                    raise ValueError("add_table requires range")
                if not _TABLE_NAME.match(name):
                    raise ValueError("Excel table name must start with a letter or underscore and contain only letters, numbers, underscore or period")
                if any(name == table.name for worksheet in book.worksheets for table in worksheet.tables.values()):
                    raise ValueError(f"Excel table already exists: {name}")
                table = Table(displayName=name, ref=ref)
                table.tableStyleInfo = TableStyleInfo(
                    name=str(operation.get("style") or "TableStyleMedium2"),
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                sheet.add_table(table)
            else:
                raise ValueError(f"Unsupported XLSX edit action: {action}")
        book.save(path)
    finally:
        book.close()


def _sheet(book, requested: Any):
    name = str(requested or "").strip()
    if not name:
        return book.active
    if name not in book.sheetnames:
        raise ValueError(f"Worksheet not found: {name}")
    return book[name]


def _sheet_name(value: Any) -> str:
    name = str(value or "").strip()
    if not name:
        raise ValueError("Worksheet name is required")
    return name[:31]


def _cell(operation: dict[str, Any], key: str = "cell") -> str:
    cell = str(operation.get(key, "")).strip().upper()
    if not cell:
        raise ValueError(f"{key} requires a cell address")
    return cell


def _hex(value: Any) -> str:
    text = str(value or "").strip().lstrip("#").upper()
    if not text:
        return ""
    if len(text) not in {6, 8} or any(character not in "0123456789ABCDEF" for character in text):
        raise ValueError("Colour must be a 6 or 8 digit hexadecimal value")
    return text


def _font_rgb(cell) -> str:
    if cell.font.color and cell.font.color.type == "rgb" and cell.font.color.rgb:
        return str(cell.font.color.rgb)
    return ""
